from io import BytesIO
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from addons.web.controllers.main import ExcelExport
from odoo import models, fields, api
from odoo import http, tools
from odoo.http import content_disposition, dispatch_rpc, serialize_exception as _serialize_exception, Response, request


class Warehouse(models.Model):
    _name = 'sharevan.warehouse'
    MODEL = 'sharevan.warehouse'

    # _inherit = 'stock.warehouse'
    _description = 'warehouse'
    _inherit = 'sharevan.warehouse'

    areaInfo = fields.Char(string='areaInfo', store=False)

    def export_warehouse(self):
        field_names = ["name", "warehouse_code", "phone", "state_id"]
        excel_headers = ["Warehouse name", "Warehouse code", "Phone", "Province"]
        excel_headers.insert(0, "No.")
        data = self.search(args=[('name', 'ilike', 'L1')])
        xp = ExcelExport()

        export_data = data.export_data(field_names).get('datas', [])
        index = 1
        for data in export_data:
            data.prepend(index)
            index += 1
        response_data = xp.from_data(excel_headers, export_data)
        return request.make_response(data=response_data,
                                     headers=[('Content-Disposition',
                                               content_disposition(self._name.replace('.', '_'))),
                                              ('Content-Type',
                                               'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')],
                                     cookies={'fileToken': 'dummytoken'})

    def get_template(self):
        template_name = "sharevan_warehouse.xlsx"
        directory = "mymodule/share_van_order/static/xlsx/"
        wb = Workbook()
        ws_report = wb.active
        ws_report.title = "Warehouse report"
        ws_validation = wb.create_sheet("Validation")
        # wb.active = 1
        ws_report['A1'] = "aslk aslkg a"

        dv = DataValidation(type="list", formula1='"Dog,Cat,Bat"', allow_blank=True)
        dv.add('B1:B1048576')
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return request.make_response(data=out,
                                     headers=[('Content-Disposition',
                                               content_disposition(self._name.replace('.', '_'))),
                                              ('Content-Type',
                                               'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')],
                                     cookies={'fileToken': 'dummytoken'})


class ShareVanChat(models.Model):
    _description = "Share Van Chat"
    _name = 'sharevan.temp'
    _inherit = 'sharevan.temp'
